#!/usr/bin/env python

import rospy
import math
from visualization_msgs.msg import Marker
import os
import simpleaudio as sa
import threading
import random
import numpy as np
from gazebo_msgs.srv import SetModelState
from gazebo_msgs.msg import ModelState
from sensor_msgs.msg import JoyFeedbackArray,JoyFeedback
import subprocess
import time
from geometry_msgs.msg import Twist, Point
from openpyxl import Workbook, load_workbook # type: ignore

markers = Marker()
def goal_sphere(config):
    
    global markers
    # 创建Marker消息

    markers.header.frame_id = "odom"  # 设置参考坐标系
    markers.header.stamp = rospy.Time.now()
    markers.ns = "sphere"  # 命名空间
    markers.id = 2  # ID
    markers.type = Marker.SPHERE  # 类型设置为球体
    markers.action = Marker.ADD  # 操作类型

    # 设置球体的中心坐标和大小
    markers.pose.position.x = config.goalX  # 圆心 x 坐标
    markers.pose.position.y = config.goalY  # 圆心 y 坐标
    markers.pose.position.z = 0.0  # 圆心 z 坐标
    markers.pose.orientation.x = 0.0
    markers.pose.orientation.y = 0.0
    markers.pose.orientation.z = 0.0
    markers.pose.orientation.w = 1.0

    markers.scale.x = config.robot_radius*4  # 球体在 x 方向的大小
    markers.scale.y = config.robot_radius*4  # 球体在 y 方向的大小
    markers.scale.z = config.robot_radius*4  # 球体在 z 方向的大小

    # 设置颜色
    markers.color.r = 0.0  # 红色分量
    markers.color.g = 1.0  # 绿色分量
    markers.color.b = 0.0  # 蓝色分量
    markers.color.a = 1.0  # 透明度

def get_time(start_time):

    end_time = rospy.get_time()  # 获取结束时间
    elapsed_time = end_time - start_time  # 计算运行时间
    print("spend time: %.2f 秒" % elapsed_time)
    return elapsed_time

def odom_callback(config):
    delta_distance = math.sqrt((config.x - config.prev_x)**2 + (config.y - config.prev_y)**2)
    if delta_distance<0.0001:
        delta_distance=0.0
    config.distance += delta_distance
    config.prev_x = config.x
    config.prev_y = config.y
    # 更新之前的位置
    # print("当前prev: %.2f , %.2f", config.prev_x,config.prev_y)
    # print("当前总里程: %.2f 米", config.distance)

def play_celebration_sound():
    """
    播放一个到达终点的庆祝音效（异步，不阻塞主线程）
    """
    def _play():
        notes = [
            (660, 0.3),  # E5
            (0,   0.1),  # pause
            (880, 0.3),  # A5
            (0,   0.1),
            (660, 0.5),  # E5 (长音)
        ]
        sample_rate = 44100
        for freq, dur in notes:
            if freq == 0:
                time.sleep(dur)
                continue
            t = np.linspace(0, dur, int(sample_rate * dur), False)
            wave = np.sin(freq * 2 * np.pi * t)
            audio = (wave * 32767).astype(np.int16)
            sa.play_buffer(audio, 1, 2, sample_rate).wait_done()
    
    threading.Thread(target=_play, daemon=True).start()

def read_last_value_from_column():
    # 读取Excel文件
    workbook = load_workbook("/home/frienkie/result/data_real.xlsx") # 替换为你的Excel文件名
    sheet = workbook['Sheet1']  # 替换为你的工作表名
    row=1
    for cell in sheet['P']:
            if cell.value is not None:
                row += 1
            else:
                break
    if row>1:
        last_row=row-1
    else:
        last_row=1
    # 获取第16列的最后一行的值
    valid_values = sheet.cell(row=last_row, column=16).value
    # 检查是否是数字
    if not isinstance(valid_values, (int, float)):
        return 1
    return valid_values+1

rosbag_process = None

def start_rosbag():
    """
    启动 rosbag 记录。
    :param record_topics: 要记录的 ROS 话题列表，字符串或列表形式
    :param output_file: rosbag 保存的文件路径，不需要后缀名
    """
    count=read_last_value_from_column()
    record_topics = ["/cmd_vel", "/cmd_vel_human", "/odom", "/min_d"]  # 话题

    rosbag_dir = os.path.expanduser("~/rosbag_real")  # Expands to /home/user/rosbag
    os.makedirs(rosbag_dir, exist_ok=True)  # Ensure directory exists

    # Output file path
    output_file = os.path.join(rosbag_dir, f"rosbag{count}")

    global rosbag_process
    # 构建 rosbag record 命令
    cmd = ["rosbag", "record", "-O", output_file]
    if isinstance(record_topics, list):
        cmd.extend(record_topics)
    elif isinstance(record_topics, str):
        cmd.append(record_topics)

    # 启动 rosbag 记录
    print(f"Starting rosbag recording: {' '.join(cmd)}")
    rosbag_process = subprocess.Popen(cmd)
    print("rosbag recording started.")
    return count

def stop_rosbag():
    """
    停止 rosbag 记录。
    """
    global rosbag_process
    if rosbag_process is not None:
        print("Stopping rosbag recording...")
        rosbag_process.terminate()  # 发送 SIGTERM 信号
        rosbag_process.wait()       # 等待进程结束
        rosbag_process = None
        print("rosbag recording stopped.")
    else:
        print("No rosbag recording process found.")



def save(time,distance,n,m):# n is direct switch,m is para
    file_name = "/home/frienkie/result/data_real.xlsx"

    # 检查文件是否存在
    if os.path.exists(file_name):
        # 如果文件存在，加载工作簿
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        # 如果文件不存在，创建新的工作簿
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"
    row=1
    # 查找第一列中最后一行的行号

    # 需要写入的浮点数据（这里替换为你的数据）
    data1 = get_time(time)  # 示例浮点数
    data2 = distance
    # 格式化浮点数，保留两位小数
    formatted_data1 = round(data1, 2)
    formatted_data2 = round(data2, 2)
    # 将数据写入第一列的下一行
    if n==0:
        for cell in sheet['N']:
            if cell.value is not None:
                row += 1
            else:
                break
        sheet.cell(row=row, column=1, value=formatted_data1)
        sheet.cell(row=row, column=2, value=formatted_data2)
        sheet.cell(row=row, column=14, value=4)
        sheet.cell(row=row, column=16, value=row-1)
    else:
        for cell in sheet['N']:
            if cell.value is not None:
                row += 1
            else:
                break
        sheet.cell(row=row, column=7, value=formatted_data1)
        sheet.cell(row=row, column=8, value=formatted_data2)
        sheet.cell(row=row, column=14, value=m)
        sheet.cell(row=row, column=16, value=row-1)
    # 保存工作簿
    workbook.save(file_name)