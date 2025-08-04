#!/usr/bin/env python3

import rclpy
from rclpy.node import Node
import math
from visualization_msgs.msg import Marker
from openpyxl import Workbook, load_workbook # type: ignore
import os
import simpleaudio as sa
import threading
from gazebo_model_collision_plugin.msg import Contact
import random
import numpy as np
from gazebo_msgs.srv import SetModelState
from gazebo_msgs.msg import ModelState
from sensor_msgs.msg import JoyFeedbackArray,JoyFeedback
import subprocess
import time
from geometry_msgs.msg import Twist, Point
from rclpy.qos import QoSProfile, ReliabilityPolicy, HistoryPolicy

markers = Marker()

def goal_sphere(config):
    global markers
    # 创建Marker消息
    markers.header.frame_id = "odom"  # 设置参考坐标系
    markers.header.stamp = rclpy.time.Time().to_msg()
    markers.ns = "sphere"  # 命名空间
    markers.id = 2  # ID
    markers.type = Marker.SPHERE  # 类型设置为球体
    markers.action = Marker.ADD  # 操作类型

    # 设置球体的中心坐标和大小
    markers.pose.position.x = config.goalX  # 圆心 x 坐标
    markers.pose.position.y = config.goalY  # 圆心 y 坐标
    markers.pose.position.z = 0.0  # 圆心 z 坐标
    markers.pose.orientation.x = 0.0
    markers.pose.orientation.y = 0.0
    markers.pose.orientation.z = 0.0
    markers.pose.orientation.w = 1.0

    markers.scale.x = config.robot_radius*4  # 球体在 x 方向的大小
    markers.scale.y = config.robot_radius*4  # 球体在 y 方向的大小
    markers.scale.z = config.robot_radius*4  # 球体在 z 方向的大小

    # 设置颜色
    markers.color.r = 0.0  # 红色分量
    markers.color.g = 1.0  # 绿色分量
    markers.color.b = 0.0  # 蓝色分量
    markers.color.a = 1.0  # 透明度

def get_time(start_time):
    end_time = time.time()  # 获取结束时间
    elapsed_time = end_time - start_time  # 计算运行时间
    print("spend time: %.2f 秒" % elapsed_time)
    return elapsed_time

def odom_callback(config):
    delta_distance = math.sqrt((config.x - config.prev_x)**2 + (config.y - config.prev_y)**2)
    if delta_distance<0.0001:
        delta_distance=0.0
    config.distance += delta_distance
    config.prev_x = config.x
    config.prev_y = config.y

def play_celebration_sound():
    """
    播放一个到达终点的庆祝音效（异步，不阻塞主线程）
    """
    def _play():
        notes = [
            (660, 0.3),  # E5
            (0,   0.1),  # pause
            (880, 0.3),  # A5
            (0,   0.1),
            (660, 0.5),  # E5 (长音)
        ]
        sample_rate = 44100
        for freq, dur in notes:
            if freq == 0:
                time.sleep(dur)
                continue
            t = np.linspace(0, dur, int(sample_rate * dur), False)
            wave = np.sin(freq * 2 * np.pi * t)
            audio = (wave * 32767).astype(np.int16)
            sa.play_buffer(audio, 1, 2, sample_rate).wait_done()
    
    threading.Thread(target=_play, daemon=True).start()

def save(time_val, distance, count_time, n, m, chizu):  # n is direct switch,m is para
    file_name = "/home/frienkie/result/data.xlsx"

    # 检查文件是否存在
    if os.path.exists(file_name):
        # 如果文件存在，加载工作簿
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        # 如果文件不存在，创建新的工作簿
        workbook = Workbook()
        sheet = workbook.active
        # 添加标题行
        sheet.append(['Time', 'Distance', 'Count', 'Direct', 'Param', 'Map'])

    # 添加数据行
    sheet.append([time_val, distance, count_time, n, m, chizu])

    # 保存文件
    workbook.save(file_name)
    print("Data saved to Excel file")

class StringMessageCounter:
    def __init__(self):
        # 设置监听的时间间隔
        self.count_time = 0.0
        self.send_count = 0
        self.last_time = time.time()
        self.interval = 1.0  # 1秒间隔

    def callbackobs(self, msg):
        current_time = time.time()
        if current_time - self.last_time >= self.interval:
            self.count_time += 1.0
            self.send_count += 1
            self.last_time = current_time

def set_robot_position(model_name, position, orientation):
    """
    Set robot position in Gazebo using ROS2 service
    """
    try:
        # Create a node for the service call
        node = Node('set_robot_position_node')
        
        # Create service client
        client = node.create_client(SetModelState, '/gazebo/set_model_state')
        
        # Wait for service to be available
        while not client.wait_for_service(timeout_sec=1.0):
            node.get_logger().info('Waiting for /gazebo/set_model_state service...')
        
        # Create request
        request = SetModelState.Request()
        request.model_state.model_name = model_name
        request.model_state.pose.position.x = position[0]
        request.model_state.pose.position.y = position[1]
        request.model_state.pose.position.z = position[2]
        request.model_state.pose.orientation.x = orientation[0]
        request.model_state.pose.orientation.y = orientation[1]
        request.model_state.pose.orientation.z = orientation[2]
        request.model_state.pose.orientation.w = orientation[3]
        
        # Call service
        future = client.call_async(request)
        rclpy.spin_until_future_complete(node, future)
        
        if future.result() is not None:
            node.get_logger().info('Robot position set successfully')
        else:
            node.get_logger().error('Failed to set robot position')
            
    except Exception as e:
        print(f"Error setting robot position: {e}")
    finally:
        if 'node' in locals():
            node.destroy_node()

def read_last_value_from_column():
    # 读取Excel文件
    file_name = "/home/frienkie/result/data.xlsx"
    
    if not os.path.exists(file_name):
        return 0
    
    try:
        workbook = load_workbook(file_name)
        sheet = workbook.active
        
        # 获取最后一行的值（假设第一列是行号）
        last_row = sheet.max_row
        if last_row > 1:  # 如果有数据行
            last_value = sheet.cell(row=last_row, column=1).value
            return int(last_value) if last_value is not None else 0
        else:
            return 0
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return 0

def start_rosbag():
    """
    Start rosbag recording using ROS2
    """
    try:
        # Get current timestamp for filename
        timestamp = int(time.time())
        filename = f"/home/frienkie/rosbag/recording_{timestamp}"
        
        # Start rosbag2 recording
        cmd = f"ros2 bag record -o {filename} /odom /scan /cmd_vel /cmd_vel_human /min_d"
        process = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        
        print(f"Started rosbag recording: {filename}")
        return timestamp
        
    except Exception as e:
        print(f"Error starting rosbag: {e}")
        return int(time.time())

def stop_rosbag():
    """
    Stop rosbag recording
    """
    try:
        # Kill rosbag2 process
        subprocess.run("pkill -f 'ros2 bag record'", shell=True)
        print("Stopped rosbag recording")
    except Exception as e:
        print(f"Error stopping rosbag: {e}") 