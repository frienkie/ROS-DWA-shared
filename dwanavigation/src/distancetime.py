#!/usr/bin/env python

import rospy
import math
from visualization_msgs.msg import Marker
from openpyxl import Workbook, load_workbook # type: ignore
import os
from gazebo_model_collision_plugin.msg import Contact
import random
from gazebo_msgs.srv import SetModelState
from gazebo_msgs.msg import ModelState
import subprocess
markers = Marker()
def goal_sphere(config):
    
    global markers
    # 创建Marker消息

    markers.header.frame_id = "odom"  # 设置参考坐标系
    markers.header.stamp = rospy.Time.now()
    markers.ns = "sphere"  # 命名空间
    markers.id = 2  # ID
    markers.type = Marker.SPHERE  # 类型设置为球体
    markers.action = Marker.ADD  # 操作类型

    # 设置球体的中心坐标和大小
    markers.pose.position.x = config.goalX  # 圆心 x 坐标
    markers.pose.position.y = config.goalY  # 圆心 y 坐标
    markers.pose.position.z = 0.0  # 圆心 z 坐标
    markers.pose.orientation.x = 0.0
    markers.pose.orientation.y = 0.0
    markers.pose.orientation.z = 0.0
    markers.pose.orientation.w = 1.0

    markers.scale.x = config.robot_radius*4  # 球体在 x 方向的大小
    markers.scale.y = config.robot_radius*4  # 球体在 y 方向的大小
    markers.scale.z = config.robot_radius*4  # 球体在 z 方向的大小

    # 设置颜色
    markers.color.r = 0.0  # 红色分量
    markers.color.g = 1.0  # 绿色分量
    markers.color.b = 0.0  # 蓝色分量
    markers.color.a = 1.0  # 透明度

def get_time(start_time):

    end_time = rospy.get_time()  # 获取结束时间
    elapsed_time = end_time - start_time  # 计算运行时间
    print("spend time: %.2f 秒" % elapsed_time)
    return elapsed_time


def odom_callback(config):


    # 计算当前位置和之前位置的距离
    delta_distance = math.sqrt((config.x - config.prev_x)**2 + (config.y - config.prev_y)**2)
    config.distance += delta_distance

    # 更新之前的位置
    config.prev_x = config.x
    config.prev_y = config.y
    #print("当前总里程: %.2f 米", config.distance)

def save(time,distance,count,m,chizu):
    file_name = "/home/frienkie/result/data.xlsx"

    # 检查文件是否存在
    if os.path.exists(file_name):
        # 如果文件存在，加载工作簿
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:
        # 如果文件不存在，创建新的工作簿
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sheet1"

    # 查找第一列中最后一行的行号
    row = 1
    for cell in sheet['N']:
        if cell.value is not None:
            row += 1
        else:
            break
    # 需要写入的浮点数据（这里替换为你的数据）
    data1 = time  # 示例浮点数
    data2 = distance
    data3 = count
    # 格式化浮点数，保留两位小数
    formatted_data1 = round(data1, 2)
    formatted_data2 = round(data2, 2)
    # 将数据写入第一列的下一行
    sheet.cell(row=row, column=4, value=formatted_data1)
    sheet.cell(row=row, column=5, value=formatted_data2)
    sheet.cell(row=row, column=6, value=data3)
    sheet.cell(row=row, column=14, value=m)
    sheet.cell(row=row, column=15, value=chizu)
    sheet.cell(row=row, column=16, value=row-1)
    # 保存工作簿
    workbook.save(file_name)

    
class StringMessageCounter:
    def __init__(self):
        # 设置监听的时间间隔
        self.inactive_threshold = 2.0  # 秒
        self.last_message_time = None
        self.send_count = 0
        rospy.Subscriber("/gazebo/base_collision",Contact,self.callbackobs,queue_size=10)


    def callbackobs(self,msg):
        current_time = rospy.get_time()
        
        # 判断是否是新的一次发送
        if self.last_message_time is None or \
           (current_time - self.last_message_time) > self.inactive_threshold:
            self.send_count += 1
            print(f"New message batch detected. Total count: {self.send_count}")
        
        # 更新最后消息接收时间
        self.last_message_time = current_time



def set_robot_position(model_name, position, orientation):
    """
    瞬间移动机器人到指定坐标
    :param model_name: 模型名称 (例如 "robot")
    :param position: 坐标 [x, y, z]
    :param orientation: 四元数 [x, y, z, w]
    """
    # 等待服务
    rospy.wait_for_service('/gazebo/set_model_state')
    
    try:
        # 创建服务代理
        set_state = rospy.ServiceProxy('/gazebo/set_model_state', SetModelState)
        
        # 定义模型状态
        state_msg = ModelState()
        state_msg.model_name = model_name
        state_msg.pose.position.x = position[0]
        state_msg.pose.position.y = position[1]
        state_msg.pose.position.z = position[2]
        state_msg.pose.orientation.x = orientation[0]
        state_msg.pose.orientation.y = orientation[1]
        state_msg.pose.orientation.z = orientation[2]
        state_msg.pose.orientation.w = orientation[3]
        state_msg.twist.linear.x = 0.0
        state_msg.twist.linear.y = 0.0
        state_msg.twist.linear.z = 0.0
        state_msg.twist.angular.x = 0.0
        state_msg.twist.angular.y = 0.0
        state_msg.twist.angular.z = 0.0
        # 调用服务
        resp = set_state(state_msg)
        if resp.success:
            rospy.loginfo(f"Successfully moved {model_name} to {position}")
        else:
            rospy.logwarn(f"Failed to move {model_name}: {resp.status_message}")
    except rospy.ServiceException as e:
        rospy.logerr(f"Service call failed: {e}")


rosbag_process = None

def start_rosbag():
    """
    启动 rosbag 记录。
    :param record_topics: 要记录的 ROS 话题列表，字符串或列表形式
    :param output_file: rosbag 保存的文件路径，不需要后缀名
    """
    counter_file = "counter.txt"

    # 检查文件是否存在
    if not os.path.exists(counter_file):
        # 如果文件不存在，初始化计数器为0
        with open(counter_file, "w") as file:
            file.write("0")

    # 读取文件中的计数器值
    with open(counter_file, "r") as file:
        count = int(file.read().strip())

    # 增加计数器
    count += 1

    # 将新的计数器值写回文件
    with open(counter_file, "w") as file:
        file.write(str(count))
    
    record_topics = ["/cmd_vel", "/cmd_vel_human"]  # 示例话题
    output_file = "rosbag"
    output_file = output_file +str(count)
    global rosbag_process
    # 构建 rosbag record 命令
    cmd = ["rosbag", "record", "-O", output_file]
    if isinstance(record_topics, list):
        cmd.extend(record_topics)
    elif isinstance(record_topics, str):
        cmd.append(record_topics)

    # 启动 rosbag 记录
    print(f"Starting rosbag recording: {' '.join(cmd)}")
    rosbag_process = subprocess.Popen(cmd)
    print("rosbag recording started.")
    return count

def stop_rosbag():
    """
    停止 rosbag 记录。
    """
    global rosbag_process
    if rosbag_process is not None:
        print("Stopping rosbag recording...")
        rosbag_process.terminate()  # 发送 SIGTERM 信号
        rosbag_process.wait()       # 等待进程结束
        rosbag_process = None
        print("rosbag recording stopped.")
    else:
        print("No rosbag recording process found.")